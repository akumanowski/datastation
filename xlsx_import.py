#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------------
# Created By  : akumanowski
# Created Date: 25/05/2023
# version ='1.0'
# ---------------------------------------------------------------------------
"""
Модуль консольного приложения DataStation, реализующий чтение данных из xlsx-файлов
"""
# ---------------------------------------------------------------------------
import datetime
from typing import List

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# создаем класс для чтения xlsx файлов
class DataReader:
    def __init__(self, filename: str, date: datetime.date):
        self.workbook = load_workbook(filename=filename)
        self.sheet = self.workbook.active
        self.max_rows = self.sheet.max_row
        self.date = date
        self.db_fields = [
            'sid', 'company', 'fact_qliq_1', 'fact_qliq_2', 'fact_qoil_1', 'fact_qoil_2',
            'forecast_qliq_1', 'forecast_qliq_2', 'forecast_qoil_1', 'forecast_qoil_2', 'date'
        ]

    def read_xlsx(self) -> List:
        data_stream = []
        for row in range(4, self.max_rows + 1):
            record = {}
            for col in range(len(self.db_fields) - 1):
                field_value = self.sheet.cell(row=row, column=col + 1).value
                record[self.db_fields[col]] = field_value
            record['date'] = self.date
            data_stream.append(record)
        return data_stream
